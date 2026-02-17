"""
Einmaliges Import-Skript: Liest die bestehende Excel-Datei
und fuellt die SQLite-Datenbank mit den Kontaktdaten.

Aufruf:
    python3 import_data.py
"""

import os
import sys
from datetime import date

import openpyxl

# Pfad zur Excel-Datei
EXCEL_PATH = os.path.join(
    os.path.expanduser("~"),
    "Desktop", "Büro", "Tierschutz Konzert", "Datenbank",
    "Partner_und_Kontakte_vorausgefuellt.xlsx",
)


def import_contacts():
    from app import create_app
    from models import db, Contact, EMOJI_TO_STATUS

    app = create_app()

    with app.app_context():
        # Pruefen ob bereits Daten vorhanden
        existing = Contact.query.count()
        if existing > 0:
            print(f"Datenbank enthaelt bereits {existing} Kontakte.")
            if "--force" not in sys.argv:
                print("Verwende --force um trotzdem zu importieren (bestehende Daten bleiben erhalten).")
                return
            print("--force erkannt, importiere trotzdem...")

        if not os.path.exists(EXCEL_PATH):
            print(f"Excel-Datei nicht gefunden: {EXCEL_PATH}")
            return

        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb["Partner & Kontakte"]

        imported = 0
        skipped = 0

        for row_idx in range(2, ws.max_row + 1):
            cells = [cell.value for cell in ws[row_idx]]

            kategorie = cells[0]

            # Leere Zeilen und Footer ueberspringen
            if kategorie is None:
                skipped += 1
                continue
            if "Diese Liste dient" in str(kategorie):
                skipped += 1
                continue

            # Status-Emoji -> interner Code
            raw_status = str(cells[4]).strip() if cells[4] else ""
            status = EMOJI_TO_STATUS.get(raw_status, "gelb")

            # Datum parsen (falls vorhanden)
            letzter_kontakt = None
            if cells[7] is not None:
                if isinstance(cells[7], date):
                    letzter_kontakt = cells[7]

            contact = Contact(
                kategorie=str(kategorie).strip(),
                name_organisation=str(cells[1]).strip() if cells[1] else "Unbekannt",
                kontaktperson=str(cells[2]).strip() if cells[2] else None,
                rolle=str(cells[3]).strip() if cells[3] else None,
                status=status,
                kontaktweg=str(cells[5]).strip() if cells[5] else None,
                kontaktdaten=str(cells[6]).strip() if cells[6] else None,
                letzter_kontakt=letzter_kontakt,
                naechster_schritt=str(cells[8]).strip() if cells[8] else None,
                zustaendig=str(cells[9]).strip() if cells[9] else "Daniel",
                notizen=str(cells[10]).strip() if cells[10] else None,
            )

            db.session.add(contact)
            imported += 1

        db.session.commit()
        print(f"Import abgeschlossen: {imported} Kontakte importiert, {skipped} Zeilen uebersprungen.")


if __name__ == "__main__":
    import_contacts()
